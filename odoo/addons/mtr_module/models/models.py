# -*- coding: utf-8 -*-
import base64
import csv
import datetime
import io
import json
import re
import urllib.error
import urllib.request
import uuid

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError


def _normalize_header(value):
    text = "" if value is None else str(value)
    normalized = re.sub(r"[^a-z0-9]+", "_", text.strip().lower())
    return normalized.strip("_")


def _to_float(value):
    if value in (None, ""):
        return False
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "")
    if not cleaned:
        return False
    try:
        return float(cleaned)
    except ValueError:
        return False


def _to_date(value):
    if value in (None, "", False):
        return False
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = str(value).strip()
    # Extract a date token from strings that include time or extra text.
    match = re.search(
        r"(\\d{4}-\\d{2}-\\d{2}|\\d{4}/\\d{2}/\\d{2}|\\d{1,2}/\\d{1,2}/\\d{4}|\\d{1,2}-\\d{1,2}-\\d{4})",
        text,
    )
    if match:
        text = match.group(0)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m/%d/%y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return False


def _to_int(value):
    number = _to_float(value)
    if number is False:
        return False
    return int(number)


def _clean_text(value):
    if value in (None, "", False):
        return False
    text = str(value).strip()
    return text or False


def _sanitize_headers(headers):
    seen = {}
    cleaned = []
    for idx, header in enumerate(headers, start=1):
        text = "" if header is None else str(header).strip()
        if not text:
            text = "column_%s" % idx
        base = text
        count = seen.get(base, 0)
        if count:
            text = "%s_%s" % (base, count + 1)
        seen[base] = count + 1
        cleaned.append(text)
    return cleaned


def _extract_certificate_number_from_filename(file_name):
    name = _clean_text(file_name)
    if not name:
        return False
    base = name.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    base_no_ext = re.sub(r"\.[^.]+$", "", base)
    compact = re.sub(r"[^0-9A-Za-z]+", "", base_no_ext)
    match = re.search(r"(?i)MT[0-9A-Za-z]+", compact)
    if match:
        return match.group(0)
    match = re.search(r"(?i)\b(MT[0-9A-Za-z]+)\b", base_no_ext)
    if match:
        return match.group(1)
    return False


def _normalize_certificate_number(value):
    text = _clean_text(value)
    if not text:
        return False
    base = text.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    base_no_ext = re.sub(r"\.[^.]+$", "", base)
    compact = re.sub(r"[^0-9A-Za-z]+", "", base_no_ext)
    match = re.search(r"(?i)MT[0-9A-Za-z]+", compact)
    if match:
        return match.group(0)
    match = re.search(r"(?i)\b(MT[0-9A-Za-z]+)\b", base_no_ext)
    if match:
        return match.group(1)
    return base_no_ext or text


def _normalize_key_part(value):
    text = _clean_text(value)
    if not text:
        return False
    # Normalize for stable matching: trim + uppercase + remove whitespace
    return re.sub(r"\\s+", "", text).upper()


def _make_heat_lot_key(heat, lot):
    heat_key = _normalize_key_part(heat)
    lot_key = _normalize_key_part(lot)
    if not (heat_key and lot_key):
        return False
    return "%s|%s" % (heat_key, lot_key)


_DEFAULT_N8N_TEST_WEBHOOK = "https://innovation.eoxs.com/webhook/mtr-local"

def _normalize_base64(value):
    if isinstance(value, bytes):
        return value.decode("utf-8")
    return value or ""

def _build_n8n_payload(env, file_name, file_content_base64, mtr_id=None, attachment_id=None):
    db_name = env.cr.dbname
    payload = {
        "source": "odoo13_mtr_module",
        "database": db_name,
        # Compatibility key for older n8n flows.
        "db_name": db_name,
        "file_name": file_name,
        "file_content_base64": _normalize_base64(file_content_base64),
        "uploaded_at": fields.Datetime.now().isoformat(),
        "uploaded_by": env.user.login,
    }
    if mtr_id:
        payload["mtr_id"] = mtr_id
    if attachment_id:
        payload["attachment_id"] = attachment_id
    return payload

def _post_payload_to_n8n(webhook_url, payload):
    data = json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        webhook_url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=45) as response:
        return response.read().decode("utf-8", errors="replace")

_INVENTORY_HEADER_MAP = {
    "sting_date": "date",
    "date": "date",
    "posting_date": "posting_date",
    "location_code": "location_code",
    "item_no": "item_no",
    "quantity": "quantity",
    "unit_of_measure_code": "unit_of_measure_code",
    "document_no": "document_no",
    "wsi_variant_code": "wsi_variant_code",
    "dimensions": "dimensions",
    "lot_no": "lot_number",
    "lot_number": "lot_number",
    "slab_no": "slab_number",
    "slab_number": "slab_number",
    "heat_no": "heat_number",
    "heat_number": "heat_number",
    "internal_bin": "internal_bin",
    "additional_notes": "additional_notes",
    "cost_amount_actual": "cost_amount_actual",
    "description_2": "description_2",
    "origin_code": "origin_code",
    "picked": "picked",
    "cutting_plan_no": "cutting_plan_no",
    "image_path": "image_path",
    "entry_type": "entry_type",
    "document_type": "document_type",
    "drawing": "drawing",
    "yield": "yield",
    "document_line_no": "document_line_no",
    "revision": "revision",
    "laser_quality": "laser_quality",
    "unitcost_cwt": "unitcost_cwt",
    "unit_cost_cwt": "unitcost_cwt",
    "piece_no": "piece_no",
    "variant_code": "variant_code",
    "description": "description",
    "return_reason_code": "return_reason_code",
    "serial_no": "serial_no",
    "package_no": "package_no",
    "invoiced_quantity": "invoiced_quantity",
    "inventory_by_location": "inventory_by_location",
    "inventory": "inventory",
    "expiration_date": "expiration_date",
    "remaining_quantity": "remaining_quantity",
    "shipped_qty_not_returned": "shipped_qty_not_returned",
    "reserved_quantity": "reserved_quantity",
    "qty_per_unit_of_measure": "qty_per_unit_of_measure",
    "sales_amount_expected": "sales_amount_expected",
    "sales_amount_actual": "sales_amount_actual",
    "cost_amount_expected": "cost_amount_expected",
    "cost_amount_non_invtbl": "cost_amount_non_invtbl",
    "item_description": "item_description",
    "cost_amount_expected_acy": "cost_amount_expected_acy",
    "cost_amount_actual_acy": "cost_amount_actual_acy",
    "completely_invoiced": "completely_invoiced",
    "cost_amount_non_invtbl_acy": "cost_amount_non_invtbl_acy",
    "assemble_to_order": "assemble_to_order",
    "drop_shipment": "drop_shipment",
    "open": "open",
    "open_flag": "open",
    "order_type": "order_type",
    "order_no": "order_no",
    "order_line_no": "order_line_no",
    "prod_order_comp_line_no": "prod_order_comp_line_no",
    "entry_no": "entry_no",
    "project_no": "project_no",
    "project_task_no": "project_task_no",
    "source_type": "source_type",
    "source_no": "source_no",
    "source_description": "source_description",
    "source_order_no": "source_order_no",
    "grade": "grade",
    "weight": "weight",
    "posting_date": "posting_date",
    "country_of_melt": "country_of_melt",
    "country_of_manufacture": "country_of_manufacture",
    "source_file": "source_file",
}


_INVENTORY_TARGET_SELECTION = [
    ("", "-- Skip --"),
    ("date", "Sting Date"),
    ("posting_date", "Posting Date"),
    ("location_code", "Location Code"),
    ("item_no", "Item No."),
    ("quantity", "Quantity"),
    ("unit_of_measure_code", "Unit of Measure Code"),
    ("document_no", "Document No."),
    ("wsi_variant_code", "WSI Variant Code"),
    ("dimensions", "Dimensions"),
    ("lot_number", "Lot No."),
    ("slab_number", "Slab No."),
    ("heat_number", "Heat No."),
    ("internal_bin", "Internal Bin"),
    ("additional_notes", "Additional Notes"),
    ("cost_amount_actual", "Cost Amount (Actual)"),
    ("description_2", "Description 2"),
    ("origin_code", "Origin Code"),
    ("picked", "Picked"),
    ("cutting_plan_no", "Cutting Plan No."),
    ("image_path", "Image Path"),
    ("entry_type", "Entry Type"),
    ("document_type", "Document Type"),
    ("drawing", "Drawing"),
    ("yield", "Yield"),
    ("document_line_no", "Document Line No."),
    ("revision", "Revision"),
    ("laser_quality", "Laser Quality"),
    ("unitcost_cwt", "UnitCost / CWT"),
    ("piece_no", "No. of pieces"),
    ("variant_code", "Variant Code"),
    ("description", "Description"),
    ("return_reason_code", "Return Reason Code"),
    ("serial_no", "Serial No."),
    ("package_no", "Package No."),
    ("invoiced_quantity", "Invoiced Quantity"),
    ("inventory_by_location", "Inventory by Location"),
    ("inventory", "Inventory"),
    ("expiration_date", "Expiration Date"),
    ("remaining_quantity", "Remaining Quantity"),
    ("shipped_qty_not_returned", "Shipped Qty. Not Returned"),
    ("reserved_quantity", "Reserved Quantity"),
    ("qty_per_unit_of_measure", "Qty. per Unit of Measure"),
    ("sales_amount_expected", "Sales Amount (Expected)"),
    ("sales_amount_actual", "Sales Amount (Actual)"),
    ("cost_amount_expected", "Cost Amount (Expected)"),
    ("cost_amount_non_invtbl", "Cost Amount (Non-Invtbl.)"),
    ("item_description", "Item Description"),
    ("cost_amount_expected_acy", "Cost Amount (Expected) (ACY)"),
    ("cost_amount_actual_acy", "Cost Amount (Actual) (ACY)"),
    ("completely_invoiced", "Completely Invoiced"),
    ("cost_amount_non_invtbl_acy", "Cost Amount (Non-Invtbl.)(ACY)"),
    ("assemble_to_order", "Assemble to Order"),
    ("drop_shipment", "Drop Shipment"),
    ("open", "Open"),
    ("order_type", "Order Type"),
    ("order_no", "Order No."),
    ("order_line_no", "Order Line No."),
    ("prod_order_comp_line_no", "Prod. Order Comp. Line No."),
    ("entry_no", "Entry No."),
    ("project_no", "Project No."),
    ("project_task_no", "Project Task No."),
    ("source_type", "Source Type"),
    ("source_no", "Source No."),
    ("source_description", "Source Description"),
    ("source_order_no", "Source Order No."),
    ("grade", "Grade"),
    ("weight", "Weight"),
    ("posting_date", "Posting Date"),
    ("country_of_melt", "Country of Melt"),
    ("country_of_manufacture", "Country of Manufacture"),
    ("source_file", "Source File"),
]


class MtrData(models.Model):
    _name = "mtr.data"
    _inherit = ["mail.thread", "mail.activity.mixin"]
    _description = "MTR Data"
    _rec_name = "certificate_number"
    _order = "id desc"
    _sql_constraints = [
        (
            "mtr_certificate_heat_batch_unique",
            "unique(certificate_number, heat_number, batch_number)",
            "An MTR with the same certificate number, heat number, and batch number already exists.",
        )
    ]

    batch_number = fields.Char(string="Batch No.", required=True, index=True)
    heat_number = fields.Char(string="Heat No.")
    piece_no = fields.Integer(string="No. of pieces")
    grade = fields.Char()
    manufacturer = fields.Char()
    certificate_number = fields.Char(string="Certificate No.", required=True, index=True)
    certificate_date = fields.Date()

    c_element = fields.Float(string="C", digits=(16, 5))
    mn_element = fields.Float(string="Mn", digits=(16, 5))
    si_element = fields.Float(string="Si", digits=(16, 5))
    p_element = fields.Float(string="P", digits=(16, 5))
    s_element = fields.Float(string="S", digits=(16, 5))
    b_element = fields.Float(string="B", digits=(16, 5))
    v_element = fields.Float(string="V", digits=(16, 5))
    nb_element = fields.Float(string="Nb", digits=(16, 5))
    ti_element = fields.Float(string="Ti", digits=(16, 5))
    al_element = fields.Float(string="Al", digits=(16, 5))
    ca_element = fields.Float(string="Ca", digits=(16, 5))
    zr_element = fields.Float(string="Zr", digits=(16, 5))
    zn_element = fields.Float(string="Zn", digits=(16, 5))
    sn_element = fields.Float(string="Sn", digits=(16, 5))
    cu_element = fields.Float(string="Cu", digits=(16, 5))
    ni_element = fields.Float(string="Ni", digits=(16, 5))
    cr_element = fields.Float(string="Cr", digits=(16, 5))
    mo_element = fields.Float(string="Mo", digits=(16, 5))
    n_element = fields.Float(string="N", digits=(16, 5))
    ce = fields.Float(
        string="CE",
        digits=(16, 5),
        compute="_compute_ce",
        store=True,
        readonly=True,
    )

    yield_strength = fields.Float()
    tensile_strength = fields.Float()
    elongation = fields.Float()
    reduction_area = fields.Float(default=0.0)
    hardness = fields.Float(default=0.0)
    thickness = fields.Float(digits=(16, 5))
    impact_charpy = fields.Float()
    plate_dimension = fields.Char()
    direction = fields.Char()

    impact_test_temp = fields.Float()
    impact_coupon_size = fields.Char()
    impact_specimen_1 = fields.Float()
    impact_specimen_2 = fields.Float()
    impact_specimen_3 = fields.Float()
    impact_average = fields.Float()

    country_of_melt = fields.Char()
    country_of_manufacture = fields.Char()

    uploaded_at = fields.Datetime(default=fields.Datetime.now)
    n8n_status = fields.Selection(
        selection=[
            ("sent", "Sent for Processing"),
            ("processed", "Processed"),
            ("failed", "Failed"),
        ],
        default="processed",
    )
    n8n_status_at = fields.Datetime()
    n8n_status_log = fields.Text()

    def _append_n8n_log(self, message):
        timestamp = fields.Datetime.now().isoformat()
        for record in self:
            entry = "[%s] %s" % (timestamp, message)
            if record.n8n_status_log:
                record.n8n_status_log = "%s\n%s" % (record.n8n_status_log, entry)
            else:
                record.n8n_status_log = entry

    @api.depends(
        "c_element",
        "mn_element",
        "cr_element",
        "mo_element",
        "v_element",
        "ni_element",
        "cu_element",
    )
    def _compute_ce(self):
        for record in self:
            c = record.c_element or 0.0
            mn = record.mn_element or 0.0
            cr = record.cr_element or 0.0
            mo = record.mo_element or 0.0
            v = record.v_element or 0.0
            ni = record.ni_element or 0.0
            cu = record.cu_element or 0.0
            record.ce = round(c + (mn / 6.0) + ((cr + mo + v) / 5.0) + ((ni + cu) / 15.0), 5)

    def _upsert_match_domain(self, certificate_number, heat_number, batch_number):
        cert = (certificate_number or "").strip()
        heat = (heat_number or "").strip()
        batch = (batch_number or "").strip()
        if cert and heat and batch:
            return [
                ("certificate_number", "=", cert),
                ("heat_number", "=", heat),
                ("batch_number", "=", batch),
            ]
        domain = []
        if cert:
            domain.append(("certificate_number", "=", cert))
        if heat:
            domain.append(("heat_number", "=", heat))
        if batch:
            domain.append(("batch_number", "=", batch))
        return domain

    def _is_pending_placeholder(self, record):
        if not record:
            return False
        record.ensure_one()
        cert = (record.certificate_number or "").strip()
        batch = (record.batch_number or "").strip()
        return cert.startswith("PENDING-") or batch.startswith("PENDING-")

    def _payload_has_real_mtr_data(self, payload):
        """
        Refresh the join report only when n8n sends actual MTR content.
        Placeholder creation/update payloads still carry identifiers, but they
        should not trigger a refresh until real values arrive.
        """
        if not isinstance(payload, dict):
            return False

        ignored_keys = {
            "mtr_id",
            "token",
            "source",
            "database",
            "db_name",
            "file_name",
            "file_content_base64",
            "attachment_id",
            "uploaded_at",
            "uploaded_by",
            "force_create",
            "create_new",
            "n8n_status",
            "n8n_status_at",
        }
        for key, value in payload.items():
            if key in ignored_keys:
                continue
            if value in (None, "", False, [], {}):
                continue
            if key in {"certificate_number", "batch_number", "heat_number"} and str(value).startswith("PENDING-"):
                continue
            return True
        return False

    def _drop_legacy_heat_constraint_if_present(self):
        self.env.cr.execute("SELECT to_regclass('public.mtr_data')")
        if not self.env.cr.fetchone()[0]:
            return
        for index_name in (
            "mtr_data_mtr_heat_unique",
            "mtr_data_mtr_heat_batch_unique",
            "mtr_data_mtr_batch_unique",
            "mtr_data_mtr_certificate_heat_batch_unique",
        ):
            self.env.cr.execute(
                "SELECT 1 FROM pg_constraint WHERE conname = %s AND conrelid = 'mtr_data'::regclass",
                (index_name,),
            )
            if self.env.cr.fetchone():
                self.env.cr.execute("ALTER TABLE mtr_data DROP CONSTRAINT IF EXISTS %s" % index_name)
                continue
            self.env.cr.execute("DROP INDEX IF EXISTS %s" % index_name)

    @api.constrains("batch_number", "certificate_number")
    def _check_required_keys(self):
        for record in self:
            if record.n8n_status == "sent" and (record.batch_number or "").startswith("PENDING-"):
                if record.batch_number and record.certificate_number:
                    continue
            if not record.batch_number:
                raise ValidationError(_("Batch Number is required."))
            if not record.certificate_number:
                raise ValidationError(_("Certificate Number is required."))

    @api.model
    def upsert_from_payload(self, payload):
        result = self._upsert_from_payload_impl(payload)
        if self._payload_has_real_mtr_data(payload):
            self.env["mtr.inventory.join.report"].sudo().refresh_view()
        return result

    def _upsert_from_payload_impl(self, payload):
        if not isinstance(payload, dict):
            raise UserError(_("Payload must be a dictionary."))

        self._drop_legacy_heat_constraint_if_present()

        mtr_id = payload.get("mtr_id")
        token = uuid.uuid4().hex[:8]
        heat_number = (payload.get("heat_number") or "").strip()
        batch_number = (payload.get("batch_number") or payload.get("heat_number") or "").strip()
        file_name = payload.get("file_name") or payload.get("source_file") or ""
        filename_certificate_number = _extract_certificate_number_from_filename(file_name)
        certificate_number = _normalize_certificate_number(filename_certificate_number or "")
        force_create = bool(payload.get("force_create") or payload.get("create_new"))
        if not mtr_id and not batch_number:
            raise UserError(_("batch_number is required."))

        values = {
            "grade": payload.get("grade"),
            "manufacturer": payload.get("manufacturer"),
            "certificate_date": _to_date(payload.get("certificate_date")),
            "piece_no": _to_int(
                payload.get("piece_no")
                or payload.get("piece_number")
                or payload.get("pieceCount")
                or payload.get("piece_count")
            ),
            "c_element": _to_float(payload.get("c_element") or payload.get("c")),
            "mn_element": _to_float(payload.get("mn_element") or payload.get("mn")),
            "si_element": _to_float(payload.get("si_element") or payload.get("si")),
            "p_element": _to_float(payload.get("p_element") or payload.get("p")),
            "s_element": _to_float(payload.get("s_element") or payload.get("s")),
            "b_element": _to_float(payload.get("b_element") or payload.get("b") or payload.get("boron")),
            "v_element": _to_float(payload.get("v_element") or payload.get("v") or payload.get("vanadium")),
            "nb_element": _to_float(payload.get("nb_element") or payload.get("nb") or payload.get("niobium")),
            "ti_element": _to_float(payload.get("ti_element") or payload.get("ti") or payload.get("titanium")),
            "al_element": _to_float(payload.get("al_element") or payload.get("al") or payload.get("aluminum") or payload.get("aluminium")),
            "ca_element": _to_float(payload.get("ca_element") or payload.get("ca") or payload.get("calcium")),
            "zr_element": _to_float(payload.get("zr_element") or payload.get("zr") or payload.get("zirconium")),
            "zn_element": _to_float(payload.get("zn_element") or payload.get("zn") or payload.get("zinc")),
            "sn_element": _to_float(payload.get("sn_element") or payload.get("sn") or payload.get("tin")),
            "cu_element": _to_float(payload.get("cu_element") or payload.get("cu")),
            "ni_element": _to_float(payload.get("ni_element") or payload.get("ni")),
            "cr_element": _to_float(payload.get("cr_element") or payload.get("cr")),
            "mo_element": _to_float(payload.get("mo_element") or payload.get("mo")),
            "n_element": _to_float(payload.get("n_element") or payload.get("n")),
            "yield_strength": _to_float(payload.get("yield_strength")),
            "tensile_strength": _to_float(payload.get("tensile_strength")),
            "elongation": _to_float(payload.get("elongation")),
            "reduction_area": _to_float(payload.get("reduction_area")) or 0.0,
            "hardness": _to_float(payload.get("hardness")) or 0.0,
            "thickness": _to_float(
                payload.get("thickness")
                or payload.get("thickness_mm")
                or payload.get("plate_thickness")
            ),
            "impact_charpy": _to_float(payload.get("impact_charpy") or payload.get("impactcharpy") or payload.get("impact_charpy_average") or payload.get("impactCharpy")),
            "plate_dimension": payload.get("plate_dimension") or payload.get("plate_dimensions") or payload.get("plate dimension") or payload.get("Plate dimension") or payload.get("dimensions"),
            "direction": payload.get("direction") or payload.get("dir") or payload.get("Dir"),
            "impact_test_temp": _to_float(payload.get("impact_test_temp")),
            "impact_coupon_size": payload.get("impact_coupon_size"),
            "impact_specimen_1": _to_float(payload.get("impact_specimen_1")),
            "impact_specimen_2": _to_float(payload.get("impact_specimen_2")),
            "impact_specimen_3": _to_float(payload.get("impact_specimen_3")),
            "impact_average": _to_float(payload.get("impact_average")),
            "country_of_melt": (
                payload.get("country_of_melt")
                or payload.get("countryOfMeltPour")
                or payload.get("countryOfMelt")
            ),
            "country_of_manufacture": (
                payload.get("country_of_manufacture")
                or payload.get("countryOfManufacture")
                or payload.get("countryOfOrigin")
            ),
            "uploaded_at": fields.Datetime.now(),
            "n8n_status": "processed",
            "n8n_status_at": fields.Datetime.now(),
        }
        if certificate_number:
            values["certificate_number"] = certificate_number
        elif not mtr_id:
            values["certificate_number"] = "PENDING-%s" % token
        if batch_number:
            values["batch_number"] = batch_number
        if heat_number:
            values["heat_number"] = heat_number

        match_domain = self._upsert_match_domain(
            values.get("certificate_number"),
            values.get("heat_number"),
            values.get("batch_number"),
        )

        with self.env.cr.savepoint():
            if mtr_id:
                target = self.search([("id", "=", mtr_id)], limit=1)
                if not target:
                    raise UserError(_("MTR record not found for id %s.") % mtr_id)
                # If the incoming composite key already exists on another record, keep that
                # real record and remove the pending placeholder.
                if match_domain:
                    existing = self.search(match_domain, limit=1)
                    if existing and existing.id != target.id:
                        if self._is_pending_placeholder(target):
                            target.sudo().unlink()
                            existing.write(values)
                            existing._append_n8n_log("Processed")
                            existing.sudo().message_post(
                                body="MTR data updated",
                                message_type="comment",
                                subtype_id=self.env.ref("mail.mt_note").id,
                            )
                            return {
                                "id": existing.id,
                                "operation": "pending_deleted_updated_existing",
                                "warning": "Pending record deleted and existing record updated.",
                            }
                        # If the target is not pending, fall through to update it.
                target.write(values)
                target._append_n8n_log("Processed")
                target.sudo().message_post(
                    body="MTR data updated",
                    message_type="comment",
                    subtype_id=self.env.ref("mail.mt_note").id,
                )
                return {"id": target.id, "operation": "updated"}

            if not force_create:
                pending = self.search(
                    [
                        ("n8n_status", "in", ["sent", "failed"]),
                        ("batch_number", "like", "PENDING-"),
                    ],
                    order="id desc",
                    limit=1,
                )
                if pending:
                    if match_domain:
                        existing = self.search(match_domain, limit=1)
                        if existing and existing.id != pending.id:
                            # Delete the placeholder pending record whenever a real record with
                            # the same composite key already exists.
                            if self._is_pending_placeholder(pending):
                                pending_id = pending.id
                                pending.sudo().unlink()
                                existing._append_n8n_log(
                                    "Duplicate certificate detected. Pending MTR %s deleted." % pending_id
                                )
                                existing.sudo().message_post(
                                    body="Duplicate certificate detected. Pending MTR %s deleted."
                                    % pending_id,
                                    message_type="comment",
                                    subtype_id=self.env.ref("mail.mt_note").id,
                                )
                                return {
                                    "id": existing.id,
                                    "operation": "pending_deleted_updated_existing",
                                    "warning": "Pending record deleted and existing record updated.",
                                }
                            existing.write(values)
                            existing._append_n8n_log("Processed")
                            existing.sudo().message_post(
                                body="MTR data updated",
                                message_type="comment",
                                subtype_id=self.env.ref("mail.mt_note").id,
                            )
                            return {"id": existing.id, "operation": "updated"}
                    pending.write(values)
                    pending._append_n8n_log("Processed")
                    pending.sudo().message_post(
                        body="MTR data updated",
                        message_type="comment",
                        subtype_id=self.env.ref("mail.mt_note").id,
                    )
                    return {"id": pending.id, "operation": "updated"}

            if force_create:
                # If a record with the same composite key already exists, update it instead of
                # raising a unique constraint error.
                existing_force = self.search(match_domain, limit=1) if match_domain else False
                if existing_force:
                    existing_force.write(values)
                    existing_force._append_n8n_log("Processed")
                    existing_force.sudo().message_post(
                        body="MTR data updated",
                        message_type="comment",
                        subtype_id=self.env.ref("mail.mt_note").id,
                    )
                    return {"id": existing_force.id, "operation": "updated"}
                created = self.with_context(mail_create_nolog=True).create(values)
                created._append_n8n_log("Processed")
                created.sudo().message_post(
                    body="MTR data updated",
                    message_type="comment",
                    subtype_id=self.env.ref("mail.mt_note").id,
                )
                return {"id": created.id, "operation": "created"}

            existing = self.search(match_domain, limit=1) if match_domain else False
            if existing:
                existing.write(values)
                existing._append_n8n_log("Processed")
                existing.sudo().message_post(
                    body="MTR data updated",
                    message_type="comment",
                    subtype_id=self.env.ref("mail.mt_note").id,
                )
                return {"id": existing.id, "operation": "updated"}
            created = self.with_context(mail_create_nolog=True).create(values)
            created._append_n8n_log("Processed")
            created.sudo().message_post(
                body="MTR data updated",
                message_type="comment",
                subtype_id=self.env.ref("mail.mt_note").id,
            )
            return {"id": created.id, "operation": "created"}

    def action_resend_to_n8n(self):
        self.ensure_one()
        webhook_url = self.env["ir.config_parameter"].sudo().get_param("mtr_module.n8n_webhook_url")
        if not webhook_url:
            raise UserError(_("Webhook URL is required."))

        attachment = self.env["ir.attachment"].search(
            [
                ("res_model", "=", "mtr.data"),
                ("res_id", "=", self.id),
                ("mimetype", "=", "application/pdf"),
            ],
            order="id desc",
            limit=1,
        )
        if not attachment or not attachment.datas:
            raise UserError(_("No PDF attachment found for this MTR."))

        encoded_file = attachment.datas
        if isinstance(encoded_file, bytes):
            encoded_file = encoded_file.decode("utf-8")

        payload = _build_n8n_payload(
            self.env,
            attachment.name or "MTR.pdf",
            encoded_file,
            mtr_id=self.id,
            attachment_id=attachment.id,
        )
        try:
            _post_payload_to_n8n(webhook_url, payload)
        except urllib.error.HTTPError as exc:
            self.write({"n8n_status": "failed", "n8n_status_at": fields.Datetime.now()})
            self._append_n8n_log("Resend failed (%s)" % exc.code)
            error_body = exc.read().decode("utf-8", errors="replace")
            raise UserError(_("Webhook failed (%s): %s") % (exc.code, error_body[:400]))
        except Exception as exc:
            self.write({"n8n_status": "failed", "n8n_status_at": fields.Datetime.now()})
            self._append_n8n_log("Resend failed (%s)" % str(exc))
            raise UserError(_("Failed to reach webhook: %s") % str(exc))

        self.write({"n8n_status": "sent", "n8n_status_at": fields.Datetime.now()})
        self._append_n8n_log("Resent for processing")

    def action_resend_pending_to_n8n(self):
        pending = self.search(
            [
                "|",
                ("heat_number", "like", "PENDING-"),
                ("n8n_status", "in", ["sent", "failed"]),
            ],
            order="id asc",
        )
        if not pending:
            raise UserError(_("No pending MTR records found."))

        failures = []
        for rec in pending:
            try:
                rec.action_resend_to_n8n()
            except Exception as exc:
                rec._append_n8n_log("Bulk resend failed (%s)" % str(exc))
                failures.append(rec.id)

        if failures:
            msg = "Resent with failures. Failed record IDs: %s" % ", ".join(map(str, failures))
            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {"title": "Resend Pending", "message": msg, "type": "warning"},
            }

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {"title": "Resend Pending", "message": "All pending records resent.", "type": "success"},
        }


class InventoryRecord(models.Model):
    _name = "inventory.record"
    _description = "Inventory Record"
    _rec_name = "lot_number"
    _order = "id desc"
    _table = "inventory"

    date = fields.Date(string="Sting Date")
    location_code = fields.Char()
    item_no = fields.Char(string="Item No.", index=True)
    quantity = fields.Float(digits=(16, 5))
    unit_of_measure_code = fields.Char(string="Unit of Measure Code")
    document_no = fields.Char(string="Document No.")
    wsi_variant_code = fields.Char()
    dimensions = fields.Char()
    lot_number = fields.Char(string="Lot No.", index=True)
    heat_number = fields.Char(string="Heat No.", index=True)
    slab_number = fields.Char(string="Slab No.")
    internal_bin = fields.Char()
    additional_notes = fields.Text()
    cost_amount_actual = fields.Float(string="Cost Amount (Actual)", digits=(16, 5))
    description_2 = fields.Char()
    origin_code = fields.Char()
    picked = fields.Char()
    cutting_plan_no = fields.Char()
    image_path = fields.Char()
    entry_type = fields.Char()
    document_type = fields.Char()
    drawing = fields.Char()
    yield_value = fields.Float(string="Yield", digits=(16, 5))
    document_line_no = fields.Integer(string="Document Line No.")
    revision = fields.Char()
    laser_quality = fields.Char()
    unitcost_cwt = fields.Float(string="UnitCost / CWT", digits=(16, 5))
    piece_no = fields.Integer(string="Piece No.")
    variant_code = fields.Char()
    description = fields.Char()
    return_reason_code = fields.Char()
    serial_no = fields.Char(string="Serial No.")
    package_no = fields.Char(string="Package No.")
    invoiced_quantity = fields.Float(digits=(16, 5))
    inventory_by_location = fields.Float(digits=(16, 5))
    inventory = fields.Float(digits=(16, 5))
    expiration_date = fields.Date()
    remaining_quantity = fields.Float(digits=(16, 5))
    shipped_qty_not_returned = fields.Float(string="Shipped Qty. Not Returned", digits=(16, 5))
    reserved_quantity = fields.Float(digits=(16, 5))
    qty_per_unit_of_measure = fields.Float(string="Qty. per Unit of Measure", digits=(16, 5))
    sales_amount_expected = fields.Float(string="Sales Amount (Expected)", digits=(16, 5))
    sales_amount_actual = fields.Float(string="Sales Amount (Actual)", digits=(16, 5))
    cost_amount_expected = fields.Float(string="Cost Amount (Expected)", digits=(16, 5))
    cost_amount_non_invtbl = fields.Float(string="Cost Amount (Non-Invtbl.)", digits=(16, 5))
    item_description = fields.Char()
    cost_amount_expected_acy = fields.Float(string="Cost Amount (Expected) (ACY)", digits=(16, 5))
    cost_amount_actual_acy = fields.Float(string="Cost Amount (Actual) (ACY)", digits=(16, 5))
    completely_invoiced = fields.Char()
    cost_amount_non_invtbl_acy = fields.Float(string="Cost Amount (Non-Invtbl.)(ACY)", digits=(16, 5))
    assemble_to_order = fields.Char()
    drop_shipment = fields.Char()
    open_flag = fields.Char(string="Open")
    order_type = fields.Char()
    order_no = fields.Char(string="Order No.")
    order_line_no = fields.Integer(string="Order Line No.")
    prod_order_comp_line_no = fields.Integer(string="Prod. Order Comp. Line No.")
    entry_no = fields.Integer(string="Entry No.", index=True)
    project_no = fields.Char()
    project_task_no = fields.Char()
    source_type = fields.Char()
    source_no = fields.Char(string="Source No.")
    source_description = fields.Char()
    source_order_no = fields.Char(string="Source Order No.")

    grade = fields.Char()
    weight = fields.Float(digits=(16, 5))
    posting_date = fields.Date()
    country_of_melt = fields.Char()
    country_of_manufacture = fields.Char()
    source_file = fields.Char()
    raw_row_data = fields.Text(help="Full BC row as JSON for unmapped columns.")
    heat_lot_key = fields.Char(index=True, compute="_compute_heat_lot_key", store=True)

    @api.depends("heat_number", "lot_number")
    def _compute_heat_lot_key(self):
        for record in self:
            record.heat_lot_key = _make_heat_lot_key(
                record.heat_number, record.lot_number
            )

    @api.model_create_multi
    def create(self, vals_list):
        """
        Enforce upsert behavior on (heat_number, lot_number).
        If a record already exists with the same heat + lot, update it instead of creating a duplicate.
        Skip dedup entirely when mtr_bulk_import context is set (used by the import wizard).
        """
        if self.env.context.get("mtr_bulk_import"):
            for vals in vals_list:
                heat = _clean_text(vals.get("heat_number"))
                lot = _clean_text(vals.get("lot_number"))
                vals["heat_lot_key"] = _make_heat_lot_key(heat, lot)
            return super(InventoryRecord, self).create(vals_list)

        records = self.browse()
        remaining = []
        for vals in vals_list:
            heat = _clean_text(vals.get("heat_number"))
            lot = _clean_text(vals.get("lot_number"))
            key = _make_heat_lot_key(heat, lot)
            if key:
                existing = self.search(
                    [("heat_lot_key", "=", key)],
                    limit=1,
                )
                if existing:
                    # Do not clear key fields when incoming values are blank
                    if not heat:
                        vals.pop("heat_number", None)
                    if not lot:
                        vals.pop("lot_number", None)
                    vals["heat_lot_key"] = key
                    existing.write(vals)
                    records |= existing
                    continue
            vals["heat_lot_key"] = key
            remaining.append(vals)

        if remaining:
            records |= super().create(remaining)
        return records


class InventoryImportWizard(models.TransientModel):
    _name = "inventory.import.wizard"
    _description = "Inventory Import Wizard"

    file_data = fields.Binary()
    file_name = fields.Char()
    delimiter = fields.Char(default=",")
    has_header = fields.Boolean(default=True)
    batch_size = fields.Integer(
        default=200000,
        help="Rows copied per flush. Larger values are faster for big files.",
    )
    line_ids = fields.One2many("inventory.import.mapping.line", "wizard_id", string="Header Mapping")

    def _build_mapping_lines(self):
        self.ensure_one()
        if not self.file_data or not self.file_name:
            return [(5, 0, 0)]

        headers = self._read_headers()
        if not headers:
            raise UserError(_("No data rows were found in the file."))

        lines = [(5, 0, 0)]
        for header in headers:
            if not header:
                continue
            normalized = _normalize_header(header)
            default_target = _INVENTORY_HEADER_MAP.get(normalized, "")
            lines.append(
                (
                    0,
                    0,
                    {
                        "source_header": header,
                        "target_key": default_target,
                    },
                )
            )
        return lines

    @api.onchange("file_data", "file_name", "has_header", "delimiter")
    def _onchange_import_file(self):
        for wizard in self:
            if wizard.file_data and wizard.file_name:
                try:
                    wizard.line_ids = wizard._build_mapping_lines()
                except UserError:
                    wizard.line_ids = [(5, 0, 0)]
            else:
                wizard.line_ids = [(5, 0, 0)]

    # Columns written by the raw-SQL bulk import path (order must match the tuple built below).
    _IMPORT_COLS = (
        "create_uid", "write_uid", "create_date", "write_date", "heat_lot_key",
        "date", "location_code", "item_no", "quantity", "unit_of_measure_code",
        "document_no", "wsi_variant_code", "dimensions", "lot_number", "heat_number",
        "slab_number", "internal_bin", "additional_notes", "cost_amount_actual",
        "description_2", "origin_code", "picked", "cutting_plan_no", "image_path",
        "entry_type", "document_type", "drawing", "yield_value", "document_line_no",
        "revision", "laser_quality", "unitcost_cwt", "piece_no", "variant_code",
        "description", "return_reason_code", "serial_no", "package_no",
        "invoiced_quantity", "inventory_by_location", "inventory", "expiration_date",
        "remaining_quantity", "shipped_qty_not_returned", "reserved_quantity",
        "qty_per_unit_of_measure", "sales_amount_expected", "sales_amount_actual",
        "cost_amount_expected", "cost_amount_non_invtbl", "item_description",
        "cost_amount_expected_acy", "cost_amount_actual_acy", "completely_invoiced",
        "cost_amount_non_invtbl_acy", "assemble_to_order", "drop_shipment", "open_flag",
        "order_type", "order_no", "order_line_no", "prod_order_comp_line_no", "entry_no",
        "project_no", "project_task_no", "source_type", "source_no", "source_description",
        "source_order_no", "grade", "weight", "posting_date", "country_of_melt",
        "country_of_manufacture", "raw_row_data",
    )

    def _get_batch_size(self):
        self.ensure_one()
        batch_size = self.batch_size or 200000
        if batch_size < 1000:
            return 1000
        if batch_size > 500000:
            return 500000
        return batch_size

    def _open_import_source(self):
        self.ensure_one()
        file_name = (self.file_name or "").lower()
        data = base64.b64decode(self.file_data)
        if file_name.endswith(".csv"):
            return self._open_csv_import_source(data)
        if file_name.endswith(".xlsx"):
            return self._open_xlsx_import_source(data)
        raise UserError(_("Only CSV and XLSX files are supported."))

    def _open_csv_import_source(self, data):
        text = data.decode("utf-8-sig")
        buffer = io.StringIO(text)
        reader = csv.reader(buffer, delimiter=self.delimiter or ",")
        if self.has_header:
            try:
                headers = _sanitize_headers(next(reader))
            except StopIteration:
                return [], iter(())
            return headers, reader
        return [], reader

    def _open_xlsx_import_source(self, data):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("openpyxl is required for XLSX imports."))

        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        if self.has_header:
            try:
                headers = _sanitize_headers(next(row_iter))
            except StopIteration:
                return [], iter(())
            return headers, row_iter
        return [], sheet.iter_rows(values_only=True)

    def action_import(self):
        self.ensure_one()
        if not self.file_data or not self.file_name:
            raise UserError(_("Please upload a file."))
        if self.has_header and not self.line_ids:
            raise UserError(_("Headers are still loading. Please wait until the mapping table appears, then click Import again."))

        selected_map = {}
        if self.line_ids:
            for line in self.line_ids:
                source_key = _normalize_header(line.source_header)
                if source_key and line.target_key:
                    selected_map[source_key] = line.target_key

        # inventory is referenced by transient spec-match rows, so clear the dependent
        # table in the same statement and let PostgreSQL handle the FK safely.
        self.env.cr.execute("TRUNCATE TABLE mtr_spec_match_result, inventory RESTART IDENTITY CASCADE")

        now_str = fields.Datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        uid = self.env.uid
        cr = self.env.cr
        batch_size = self._get_batch_size()

        headers, row_iter = self._open_import_source()

        # PostgreSQL COPY is the fastest bulk-load path — no row-by-row overhead.
        # Empty string in the CSV maps to NULL via NULL '' option.
        col_sql = ", ".join('"%s"' % c for c in self._IMPORT_COLS)
        copy_sql = "COPY inventory (%s) FROM STDIN WITH (FORMAT CSV, NULL '')" % col_sql

        def _v(x):
            """False → None so csv.writer emits '' which COPY treats as NULL."""
            return None if x is False else x

        buf = io.StringIO()
        writer = csv.writer(buf)
        chunk = 0
        total = 0
        any_rows = False

        col_map = None

        def _flush():
            nonlocal buf, writer, chunk
            buf.seek(0)
            cr._obj.copy_expert(copy_sql, buf)
            buf = io.StringIO()
            writer = csv.writer(buf)
            chunk = 0

        for row in row_iter:
            if not row or not any(row):
                continue
            any_rows = True

            if col_map is None:
                col_map = {}
                if not headers:
                    headers = ["column_%s" % (idx + 1) for idx in range(len(row))]
                for idx, k in enumerate(headers):
                    norm = _normalize_header(k)
                    col_map[idx] = selected_map.get(norm) or _INVENTORY_HEADER_MAP.get(norm, norm)

            if len(row) > len(headers):
                for idx in range(len(headers), len(row)):
                    headers.append("column_%s" % (idx + 1))
                    norm = _normalize_header(headers[idx])
                    col_map[idx] = selected_map.get(norm) or _INVENTORY_HEADER_MAP.get(norm, norm)

            normalized = {}
            for idx, v in enumerate(row):
                c = col_map[idx]
                if c not in normalized or normalized[c] in (None, "", False):
                    normalized[c] = v

            vals = self._map_inventory_row(normalized, row, skip_raw=True)

            heat = _clean_text(vals.get("heat_number"))
            lot = _clean_text(vals.get("lot_number"))

            writer.writerow([
                uid, uid, now_str, now_str,
                _make_heat_lot_key(heat, lot) or None,
                _v(vals.get("date")), _v(vals.get("location_code")), _v(vals.get("item_no")),
                _v(vals.get("quantity")), _v(vals.get("unit_of_measure_code")),
                _v(vals.get("document_no")), _v(vals.get("wsi_variant_code")),
                _v(vals.get("dimensions")), _v(vals.get("lot_number")), _v(vals.get("heat_number")),
                _v(vals.get("slab_number")), _v(vals.get("internal_bin")), _v(vals.get("additional_notes")),
                _v(vals.get("cost_amount_actual")), _v(vals.get("description_2")),
                _v(vals.get("origin_code")), _v(vals.get("picked")), _v(vals.get("cutting_plan_no")),
                _v(vals.get("image_path")), _v(vals.get("entry_type")), _v(vals.get("document_type")),
                _v(vals.get("drawing")), _v(vals.get("yield_value")), _v(vals.get("document_line_no")),
                _v(vals.get("revision")), _v(vals.get("laser_quality")), _v(vals.get("unitcost_cwt")),
                _to_int(vals.get("piece_no")),
                _v(vals.get("variant_code")), _v(vals.get("description")),
                _v(vals.get("return_reason_code")), _v(vals.get("serial_no")), _v(vals.get("package_no")),
                _v(vals.get("invoiced_quantity")), _v(vals.get("inventory_by_location")),
                _v(vals.get("inventory")), _v(vals.get("expiration_date")),
                _v(vals.get("remaining_quantity")), _v(vals.get("shipped_qty_not_returned")),
                _v(vals.get("reserved_quantity")), _v(vals.get("qty_per_unit_of_measure")),
                _v(vals.get("sales_amount_expected")), _v(vals.get("sales_amount_actual")),
                _v(vals.get("cost_amount_expected")), _v(vals.get("cost_amount_non_invtbl")),
                _v(vals.get("item_description")), _v(vals.get("cost_amount_expected_acy")),
                _v(vals.get("cost_amount_actual_acy")), _v(vals.get("completely_invoiced")),
                _v(vals.get("cost_amount_non_invtbl_acy")), _v(vals.get("assemble_to_order")),
                _v(vals.get("drop_shipment")), _v(vals.get("open_flag")), _v(vals.get("order_type")),
                _v(vals.get("order_no")), _v(vals.get("order_line_no")),
                _v(vals.get("prod_order_comp_line_no")), _v(vals.get("entry_no")),
                _v(vals.get("project_no")), _v(vals.get("project_task_no")), _v(vals.get("source_type")),
                _v(vals.get("source_no")), _v(vals.get("source_description")),
                _v(vals.get("source_order_no")), _v(vals.get("grade")), _v(vals.get("weight")),
                _v(vals.get("posting_date")), _v(vals.get("country_of_melt")),
                _v(vals.get("country_of_manufacture")),
                None,  # raw_row_data — skipped for speed; not needed without dedup
            ])
            chunk += 1
            total += 1

            if chunk >= batch_size:
                _flush()

        if not any_rows:
            raise UserError(_("No data rows were found in the file."))
        if chunk:
            _flush()

        self.env["inventory.record"].invalidate_cache()
        self.env["mtr.inventory.join.report"].sudo().refresh_view()
        return {"type": "ir.actions.act_window_close"}

    def _find_existing_inventory(self, Inventory, vals):
        """
        Try to find an existing inventory record to update instead of creating a duplicate.
        Priority:
          1) heat_number + lot_number (requested primary key)
          2) raw_row_data exact match (identical row)
          3) document_no + document_line_no + item_no + posting_date + location_code
          4) heat_number + item_no + posting_date + location_code
        """
        heat = (vals.get("heat_number") or "").strip()
        lot = (vals.get("lot_number") or "").strip()
        key = _make_heat_lot_key(heat, lot)
        if key:
            return Inventory.search([("heat_lot_key", "=", key)], limit=1)

        raw_row = vals.get("raw_row_data")
        if raw_row:
            rec = Inventory.search([("raw_row_data", "=", raw_row)], limit=1)
            if rec:
                return rec

        doc_no = (vals.get("document_no") or "").strip()
        doc_line = vals.get("document_line_no")
        item_no = (vals.get("item_no") or "").strip()
        posting_date = vals.get("posting_date")
        location = (vals.get("location_code") or "").strip()
        if doc_no and item_no and posting_date:
            domain = [
                ("document_no", "=", doc_no),
                ("item_no", "=", item_no),
                ("posting_date", "=", posting_date),
            ]
            if doc_line:
                domain.append(("document_line_no", "=", doc_line))
            if location:
                domain.append(("location_code", "=", location))
            rec = Inventory.search(domain, limit=1)
            if rec:
                return rec

        if heat and item_no and posting_date:
            domain = [
                ("heat_number", "=", heat),
                ("item_no", "=", item_no),
                ("posting_date", "=", posting_date),
            ]
            if location:
                domain.append(("location_code", "=", location))
            return Inventory.search(domain, limit=1)

        return Inventory.browse()

    def _map_inventory_row(self, normalized, row, skip_raw=False):
        date_value = _to_date(normalized.get("date"))
        return {
            "date": date_value,
            "location_code": normalized.get("location_code"),
            "item_no": normalized.get("item_no"),
            "quantity": _to_float(normalized.get("quantity")),
            "unit_of_measure_code": normalized.get("unit_of_measure_code"),
            "document_no": normalized.get("document_no"),
            "wsi_variant_code": normalized.get("wsi_variant_code"),
            "dimensions": normalized.get("dimensions"),
            "lot_number": _clean_text(normalized.get("lot_no") or normalized.get("lot_number")),
            "heat_number": _clean_text(normalized.get("heat_no") or normalized.get("heat_number")),
            "slab_number": _clean_text(normalized.get("slab_no") or normalized.get("slab_number")),
            "internal_bin": normalized.get("internal_bin"),
            "additional_notes": normalized.get("additional_notes"),
            "cost_amount_actual": _to_float(normalized.get("cost_amount_actual")),
            "description_2": normalized.get("description_2"),
            "origin_code": normalized.get("origin_code"),
            "picked": str(normalized.get("picked")) if normalized.get("picked") not in (None, "") else False,
            "cutting_plan_no": normalized.get("cutting_plan_no"),
            "image_path": normalized.get("image_path"),
            "entry_type": normalized.get("entry_type"),
            "document_type": normalized.get("document_type"),
            "drawing": normalized.get("drawing"),
            "yield_value": _to_float(normalized.get("yield")),
            "document_line_no": _to_int(normalized.get("document_line_no")),
            "revision": normalized.get("revision"),
            "laser_quality": normalized.get("laser_quality"),
            "unitcost_cwt": _to_float(normalized.get("unitcost_cwt")),
            "piece_no": normalized.get("piece_no"),
            "variant_code": normalized.get("variant_code"),
            "description": normalized.get("description"),
            "return_reason_code": normalized.get("return_reason_code"),
            "serial_no": normalized.get("serial_no"),
            "package_no": normalized.get("package_no"),
            "invoiced_quantity": _to_float(normalized.get("invoiced_quantity")),
            "inventory_by_location": _to_float(normalized.get("inventory_by_location")),
            "inventory": _to_float(normalized.get("inventory")),
            "expiration_date": _to_date(normalized.get("expiration_date")),
            "remaining_quantity": _to_float(normalized.get("remaining_quantity")),
            "shipped_qty_not_returned": _to_float(normalized.get("shipped_qty_not_returned")),
            "reserved_quantity": _to_float(normalized.get("reserved_quantity")),
            "qty_per_unit_of_measure": _to_float(normalized.get("qty_per_unit_of_measure")),
            "sales_amount_expected": _to_float(normalized.get("sales_amount_expected")),
            "sales_amount_actual": _to_float(normalized.get("sales_amount_actual")),
            "cost_amount_expected": _to_float(normalized.get("cost_amount_expected")),
            "cost_amount_non_invtbl": _to_float(normalized.get("cost_amount_non_invtbl")),
            "item_description": normalized.get("item_description"),
            "cost_amount_expected_acy": _to_float(normalized.get("cost_amount_expected_acy")),
            "cost_amount_actual_acy": _to_float(normalized.get("cost_amount_actual_acy")),
            "completely_invoiced": str(normalized.get("completely_invoiced"))
            if normalized.get("completely_invoiced") not in (None, "")
            else False,
            "cost_amount_non_invtbl_acy": _to_float(normalized.get("cost_amount_non_invtbl_acy")),
            "assemble_to_order": str(normalized.get("assemble_to_order"))
            if normalized.get("assemble_to_order") not in (None, "")
            else False,
            "drop_shipment": str(normalized.get("drop_shipment"))
            if normalized.get("drop_shipment") not in (None, "")
            else False,
            "open_flag": str(normalized.get("open")) if normalized.get("open") not in (None, "") else False,
            "order_type": normalized.get("order_type"),
            "order_no": normalized.get("order_no"),
            "order_line_no": _to_int(normalized.get("order_line_no")),
            "prod_order_comp_line_no": _to_int(normalized.get("prod_order_comp_line_no")),
            "entry_no": _to_int(normalized.get("entry_no")),
            "project_no": normalized.get("project_no"),
            "project_task_no": normalized.get("project_task_no"),
            "source_type": normalized.get("source_type"),
            "source_no": normalized.get("source_no"),
            "source_description": normalized.get("source_description"),
            "source_order_no": normalized.get("source_order_no"),
            "grade": normalized.get("grade"),
            "weight": _to_float(normalized.get("weight")),
            "posting_date": _to_date(normalized.get("posting_date")) or date_value,
            "country_of_melt": normalized.get("country_of_melt"),
            "country_of_manufacture": normalized.get("country_of_manufacture"),
            "raw_row_data": None if skip_raw else json.dumps(row, sort_keys=True, default=str),
        }

    def _read_rows(self):
        file_name = (self.file_name or "").lower()
        data = base64.b64decode(self.file_data)
        if file_name.endswith(".csv"):
            return self._read_csv(data)
        if file_name.endswith(".xlsx"):
            return self._read_xlsx(data)
        raise UserError(_("Only CSV and XLSX files are supported."))

    def _read_rows_preview(self, limit=100):
        rows = []
        for row in self._iter_rows():
            rows.append(row)
            if len(rows) >= limit:
                break
        return rows

    def _read_headers(self):
        self.ensure_one()
        file_name = (self.file_name or "").lower()
        data = base64.b64decode(self.file_data)
        if file_name.endswith(".csv"):
            return self._read_csv_headers(data)
        if file_name.endswith(".xlsx"):
            return self._read_xlsx_headers(data)
        raise UserError(_("Only CSV and XLSX files are supported."))

    def _read_csv_headers(self, data):
        text = data.decode("utf-8-sig")
        buffer = io.StringIO(text)
        if self.has_header:
            reader = csv.reader(buffer, delimiter=self.delimiter or ",")
            try:
                return _sanitize_headers(next(reader))
            except StopIteration:
                return []

        reader = csv.reader(buffer, delimiter=self.delimiter or ",")
        try:
            first_row = next(reader)
        except StopIteration:
            return []
        return ["column_%s" % (idx + 1) for idx in range(len(first_row))]

    def _read_xlsx_headers(self, data):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("openpyxl is required for XLSX imports."))

        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)

        if self.has_header:
            try:
                return _sanitize_headers(next(row_iter))
            except StopIteration:
                return []

        try:
            first_row = next(row_iter)
        except StopIteration:
            return []
        return ["column_%s" % (idx + 1) for idx in range(len(first_row))]

    def _iter_rows(self):
        file_name = (self.file_name or "").lower()
        data = base64.b64decode(self.file_data)
        if file_name.endswith(".csv"):
            return self._iter_csv(data)
        if file_name.endswith(".xlsx"):
            return self._iter_xlsx(data)
        raise UserError(_("Only CSV and XLSX files are supported."))

    def _read_csv(self, data):
        text = data.decode("utf-8-sig")
        buffer = io.StringIO(text)
        if self.has_header:
            raw_reader = csv.reader(buffer, delimiter=self.delimiter or ",")
            try:
                raw_headers = next(raw_reader)
            except StopIteration:
                return []
            headers = _sanitize_headers(raw_headers)
            reader = csv.DictReader(buffer, fieldnames=headers, delimiter=self.delimiter or ",")
            return [row for row in reader if any(row.values())]

        plain_reader = csv.reader(buffer, delimiter=self.delimiter or ",")
        rows = []
        for row in plain_reader:
            if not any(row):
                continue
            generated = {"column_%s" % (idx + 1): value for idx, value in enumerate(row)}
            rows.append(generated)
        return rows

    def _iter_csv(self, data):
        text = data.decode("utf-8-sig")
        buffer = io.StringIO(text)
        if self.has_header:
            raw_reader = csv.reader(buffer, delimiter=self.delimiter or ",")
            try:
                raw_headers = next(raw_reader)
            except StopIteration:
                return
            headers = _sanitize_headers(raw_headers)
            reader = csv.DictReader(buffer, fieldnames=headers, delimiter=self.delimiter or ",")
            for row in reader:
                if any(row.values()):
                    yield row
            return

        plain_reader = csv.reader(buffer, delimiter=self.delimiter or ",")
        for row in plain_reader:
            if not any(row):
                continue
            generated = {"column_%s" % (idx + 1): value for idx, value in enumerate(row)}
            yield generated

    def _read_xlsx(self, data):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("openpyxl is required for XLSX imports."))

        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active

        row_iter = sheet.iter_rows(values_only=True)
        payload = []

        if self.has_header:
            try:
                first_row = next(row_iter)
            except StopIteration:
                return []
            headers = _sanitize_headers(first_row)
        else:
            headers = []

        for line in row_iter if self.has_header else sheet.iter_rows(values_only=True):
            if not any(line):
                continue

            if not self.has_header and len(line) > len(headers):
                # Extend headers dynamically for headerless files
                for idx in range(len(headers), len(line)):
                    headers.append("column_%s" % (idx + 1))
                for existing in payload:
                    for idx in range(len(existing), len(headers)):
                        existing[headers[idx]] = ""

            row = {headers[idx]: (line[idx] if idx < len(line) else "") for idx in range(len(headers))}
            payload.append(row)

        return payload

    def _iter_xlsx(self, data):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError(_("openpyxl is required for XLSX imports."))

        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active

        row_iter = sheet.iter_rows(values_only=True)

        if self.has_header:
            try:
                first_row = next(row_iter)
            except StopIteration:
                return
            headers = _sanitize_headers(first_row)
        else:
            headers = []

        for line in row_iter if self.has_header else sheet.iter_rows(values_only=True):
            if not any(line):
                continue

            if not self.has_header and len(line) > len(headers):
                for idx in range(len(headers), len(line)):
                    headers.append("column_%s" % (idx + 1))

            row = {headers[idx]: (line[idx] if idx < len(line) else "") for idx in range(len(headers))}
            yield row


class InventoryImportMappingLine(models.TransientModel):
    _name = "inventory.import.mapping.line"
    _description = "Inventory Import Mapping Line"
    _order = "id"

    wizard_id = fields.Many2one("inventory.import.wizard", required=True, ondelete="cascade")
    source_header = fields.Char(readonly=True)
    target_key = fields.Selection(selection=_INVENTORY_TARGET_SELECTION, string="Map To")


class MtrPdfUploadWizard(models.TransientModel):
    _name = "mtr.pdf.upload.wizard"
    _description = "MTR PDF Upload Wizard"

    file_data = fields.Binary(required=True)
    file_name = fields.Char(required=True)
    webhook_url = fields.Char(
        required=True,
        default=lambda self: self.env["ir.config_parameter"].sudo().get_param("mtr_module.n8n_webhook_url")
        or _DEFAULT_N8N_TEST_WEBHOOK,
    )
    save_as_default = fields.Boolean(default=True)

    def action_send_to_n8n(self):
        self.ensure_one()
        if not self.file_data or not self.file_name:
            raise UserError(_("Please upload a PDF file."))
        if not (self.file_name or "").lower().endswith(".pdf"):
            raise UserError(_("Only PDF files are supported for MTR upload."))
        if not self.webhook_url:
            raise UserError(_("Webhook URL is required."))
        if self.save_as_default:
            self.env["ir.config_parameter"].sudo().set_param("mtr_module.n8n_webhook_url", self.webhook_url)

        token = uuid.uuid4().hex[:8]
        pending_values = {
            "batch_number": "PENDING-%s" % token,
            "heat_number": "",
            "certificate_number": "PENDING-%s" % token,
            "uploaded_at": fields.Datetime.now(),
            "n8n_status": "sent",
            "n8n_status_at": fields.Datetime.now(),
        }
        mtr_record = self.env["mtr.data"].with_context(mail_create_nolog=True).create(pending_values)
        mtr_record._append_n8n_log("Sent for processing")

        attachment = self.env["ir.attachment"].create(
            {
                "name": self.file_name,
                "res_model": "mtr.data",
                "res_id": mtr_record.id,
                "type": "binary",
                "datas": self.file_data,
                "mimetype": "application/pdf",
            }
        )
        mtr_record.sudo().message_post(
            body="MTR sent for processing",
            attachment_ids=[attachment.id],
            message_type="comment",
            subtype_id=self.env.ref("mail.mt_note").id,
        )
        mtr_record._append_n8n_log(
            "PDF attached: %s (attachment %s)" % (self.file_name, attachment.id)
        )

        encoded_file = self.file_data
        if isinstance(encoded_file, bytes):
            encoded_file = encoded_file.decode("utf-8")

        payload = _build_n8n_payload(
            self.env,
            self.file_name,
            encoded_file,
            mtr_id=mtr_record.id,
            attachment_id=attachment.id,
        )
        try:
            _post_payload_to_n8n(self.webhook_url, payload)
        except urllib.error.HTTPError as exc:
            mtr_record.write({"n8n_status": "failed", "n8n_status_at": fields.Datetime.now()})
            mtr_record._append_n8n_log("Failed (%s)" % exc.code)
            error_body = exc.read().decode("utf-8", errors="replace")
            raise UserError(_("Webhook failed (%s): %s") % (exc.code, error_body[:400]))
        except Exception as exc:
            mtr_record.write({"n8n_status": "failed", "n8n_status_at": fields.Datetime.now()})
            mtr_record._append_n8n_log("Failed (%s)" % str(exc))
            raise UserError(_("Failed to reach webhook: %s") % str(exc))

        return {"type": "ir.actions.act_window_close"}

class MtrPdfUploadMultiWizard(models.TransientModel):
    _name = "mtr.pdf.upload.multi.wizard"
    _description = "MTR PDF Upload Wizard (Multi)"

    attachment_ids = fields.Many2many(
        "ir.attachment",
        "mtr_pdf_upload_multi_attachment_rel",
        "wizard_id",
        "attachment_id",
        string="PDF Files",
        help="Upload one or more PDF files.",
    )
    webhook_url = fields.Char(
        required=True,
        default=lambda self: self.env["ir.config_parameter"].sudo().get_param("mtr_module.n8n_webhook_url")
        or _DEFAULT_N8N_TEST_WEBHOOK,
    )
    save_as_default = fields.Boolean(default=True)

    def _process_attachment(self, attachment):
        file_name = attachment.name or ""
        if not file_name.lower().endswith(".pdf"):
            raise UserError(_("Only PDF files are supported for MTR upload."))
        if not attachment.datas:
            raise UserError(_("Missing file data for %s.") % file_name)

        token = uuid.uuid4().hex[:8]
        pending_values = {
            "batch_number": "PENDING-%s" % token,
            "heat_number": "",
            "certificate_number": "PENDING-%s" % token,
            "uploaded_at": fields.Datetime.now(),
            "n8n_status": "sent",
            "n8n_status_at": fields.Datetime.now(),
        }
        mtr_record = self.env["mtr.data"].with_context(mail_create_nolog=True).create(pending_values)
        mtr_record._append_n8n_log("Sent for processing")

        attachment.write(
            {
                "res_model": "mtr.data",
                "res_id": mtr_record.id,
                "type": "binary",
                "mimetype": "application/pdf",
            }
        )
        mtr_record.sudo().message_post(
            body="MTR sent for processing",
            attachment_ids=[attachment.id],
            message_type="comment",
            subtype_id=self.env.ref("mail.mt_note").id,
        )
        mtr_record._append_n8n_log("PDF attached: %s (attachment %s)" % (file_name, attachment.id))

        payload = _build_n8n_payload(
            self.env,
            file_name,
            attachment.datas,
            mtr_id=mtr_record.id,
            attachment_id=attachment.id,
        )
        try:
            _post_payload_to_n8n(self.webhook_url, payload)
        except urllib.error.HTTPError as exc:
            mtr_record.write({"n8n_status": "failed", "n8n_status_at": fields.Datetime.now()})
            mtr_record._append_n8n_log("Failed (%s)" % exc.code)
            error_body = exc.read().decode("utf-8", errors="replace")
            raise UserError(_("Webhook failed (%s): %s") % (exc.code, error_body[:400]))
        except Exception as exc:
            mtr_record.write({"n8n_status": "failed", "n8n_status_at": fields.Datetime.now()})
            mtr_record._append_n8n_log("Failed (%s)" % str(exc))
            raise UserError(_("Failed to reach webhook: %s") % str(exc))

        mtr_record.write({"n8n_status": "sent", "n8n_status_at": fields.Datetime.now()})
        return mtr_record

    def action_send_to_n8n(self):
        self.ensure_one()
        if not self.attachment_ids:
            raise UserError(_("Please upload at least one PDF file."))
        if not self.webhook_url:
            raise UserError(_("Webhook URL is required."))
        if self.save_as_default:
            self.env["ir.config_parameter"].sudo().set_param("mtr_module.n8n_webhook_url", self.webhook_url)

        success_count = 0
        failed = []
        for attachment in self.attachment_ids:
            try:
                self._process_attachment(attachment)
                success_count += 1
            except urllib.error.HTTPError as exc:
                error_body = exc.read().decode("utf-8", errors="replace")
                failed.append("%s (HTTP %s: %s)" % (attachment.name, exc.code, error_body[:120]))
            except Exception as exc:
                failed.append("%s (%s)" % (attachment.name, str(exc)))

        message = _("Submitted %s PDF(s).") % success_count
        if failed:
            message = "%s %s" % (message, _("Failures: %s") % "; ".join(failed))

        return {"type": "ir.actions.act_window_close"}

class MtrFindLotWizard(models.TransientModel):
    _name = "mtr.find.lot.wizard"
    _description = "Find Lot Wizard"

    grade = fields.Char(required=True)
    material_origin = fields.Char()

    c_element = fields.Float(string="C")
    mn_element = fields.Float(string="Mn")
    si_element = fields.Float(string="Si")
    p_element = fields.Float(string="P")
    s_element = fields.Float(string="S")
    b_element = fields.Float(string="B")
    v_element = fields.Float(string="V")
    nb_element = fields.Float(string="Nb")
    ti_element = fields.Float(string="Ti")
    al_element = fields.Float(string="Al")
    ca_element = fields.Float(string="Ca")
    zr_element = fields.Float(string="Zr")
    zn_element = fields.Float(string="Zn")
    sn_element = fields.Float(string="Sn")
    cu_element = fields.Float(string="Cu")
    ni_element = fields.Float(string="Ni")
    cr_element = fields.Float(string="Cr")
    mo_element = fields.Float(string="Mo")
    n_element = fields.Float(string="N")

    yield_strength = fields.Float()
    tensile_strength = fields.Float()
    elongation = fields.Float()
    reduction_area = fields.Float()
    hardness = fields.Float()

    result_batch_number = fields.Char(readonly=True)
    result_lot_number = fields.Char(readonly=True)
    result_plate_number = fields.Char(readonly=True)

    def _spec_domain(self):
        domain = []
        for field_name in [
            "c_element",
            "mn_element",
            "si_element",
            "p_element",
            "s_element",
            "b_element",
            "v_element",
            "nb_element",
            "ti_element",
            "al_element",
            "ca_element",
            "zr_element",
            "zn_element",
            "sn_element",
            "cu_element",
            "ni_element",
            "cr_element",
            "mo_element",
            "n_element",
            "yield_strength",
            "tensile_strength",
            "elongation",
            "reduction_area",
            "hardness",
        ]:
            value = getattr(self, field_name)
            if value not in (False, None, ""):
                domain.append((field_name, "=", value))
        return domain

    def action_find_lot(self):
        self.ensure_one()
        if not self.grade:
            raise UserError(_("Please enter a grade."))

        Mtr = self.env["mtr.data"]
        mtr = Mtr.search([("grade", "=", self.grade)], limit=1)

        if not mtr:
            spec_domain = self._spec_domain()
            if not spec_domain:
                raise UserError(
                    _("No MTR found for grade. Please provide chemical and mechanical specs.")
                )
            mtr = Mtr.search(spec_domain, limit=1)
            if not mtr:
                raise UserError(_("No MTR found for the provided specifications."))

        if self.material_origin:
            origin = self.material_origin.strip().lower()
            mtr_origin = (
                (mtr.country_of_melt or "").strip().lower(),
                (mtr.country_of_manufacture or "").strip().lower(),
            )
            if origin not in mtr_origin:
                raise UserError(_("Material origin does not match the MTR record."))

        identifier = (mtr.batch_number or mtr.heat_number or "").strip()
        if not identifier:
            raise UserError(_("MTR record is missing batch or heat number."))

        Inventory = self.env["inventory.record"]
        inventory = Inventory.search([("heat_number", "=", identifier)])
        if not inventory:
            raise UserError(
                _("No inventory lot found for the extracted batch or heat number.")
            )

        if self.material_origin:
            origin = self.material_origin.strip().lower()
            matched = inventory.filtered(
                lambda rec: origin
                in (
                    (rec.country_of_melt or "").strip().lower(),
                    (rec.country_of_manufacture or "").strip().lower(),
                )
            )
            if not matched:
                raise UserError(_("Material origin does not match inventory records."))
            inventory = matched

        first = inventory[0]
        self.write(
            {
                "result_batch_number": identifier,
                "result_lot_number": first.lot_number,
                "result_plate_number": first.piece_no,
            }
        )

        return {
            "type": "ir.actions.act_window",
            "name": _("Find Lot Results"),
            "res_model": "mtr.find.lot.wizard",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }

class MtrPdfUploadQueue(models.Model):
    _name = "mtr.pdf.upload.queue"
    _description = "Deprecated MTR PDF Upload Queue"
    _auto = False

class MtrInventoryJoinReport(models.Model):
    _name = "mtr.inventory.join.report"
    _description = "MTR Inventory Join Report"
    _auto = False
    _rec_name = "inv_lot_number"
    _order = "join_priority asc, inv_posting_date desc, id desc"

    join_status = fields.Char(readonly=True)
    join_priority = fields.Integer(readonly=True)

    inv_entry_no = fields.Integer(readonly=True)
    inv_date = fields.Date(readonly=True)
    inv_location_code = fields.Char(readonly=True)
    inv_item_no = fields.Char(readonly=True)
    inv_quantity = fields.Float(digits=(16, 5), readonly=True)
    inv_unit_of_measure_code = fields.Char(readonly=True)
    inv_document_no = fields.Char(readonly=True)
    inv_wsi_variant_code = fields.Char(readonly=True)
    inv_dimensions = fields.Char(readonly=True)
    inv_lot_number = fields.Char(readonly=True)
    inv_heat_number = fields.Char(readonly=True)
    inv_slab_number = fields.Char(readonly=True)
    inv_internal_bin = fields.Char(readonly=True)
    inv_additional_notes = fields.Text(readonly=True)
    inv_cost_amount_actual = fields.Float(digits=(16, 5), readonly=True)
    inv_description_2 = fields.Char(readonly=True)
    inv_origin_code = fields.Char(readonly=True)
    inv_picked = fields.Char(readonly=True)
    inv_cutting_plan_no = fields.Char(readonly=True)
    inv_image_path = fields.Char(readonly=True)
    inv_entry_type = fields.Char(readonly=True)
    inv_document_type = fields.Char(readonly=True)
    inv_drawing = fields.Char(readonly=True)
    inv_yield_value = fields.Float(digits=(16, 5), readonly=True)
    inv_document_line_no = fields.Integer(readonly=True)
    inv_revision = fields.Char(readonly=True)
    inv_laser_quality = fields.Char(readonly=True)
    inv_unitcost_cwt = fields.Float(digits=(16, 5), readonly=True)
    inv_piece_no = fields.Char(readonly=True)
    inv_variant_code = fields.Char(readonly=True)
    inv_description = fields.Char(readonly=True)
    inv_return_reason_code = fields.Char(readonly=True)
    inv_serial_no = fields.Char(readonly=True)
    inv_package_no = fields.Char(readonly=True)
    inv_invoiced_quantity = fields.Float(digits=(16, 5), readonly=True)
    inv_inventory_by_location = fields.Float(digits=(16, 5), readonly=True)
    inv_inventory = fields.Float(digits=(16, 5), readonly=True)
    inv_expiration_date = fields.Date(readonly=True)
    inv_remaining_quantity = fields.Float(digits=(16, 5), readonly=True)
    inv_shipped_qty_not_returned = fields.Float(digits=(16, 5), readonly=True)
    inv_reserved_quantity = fields.Float(digits=(16, 5), readonly=True)
    inv_qty_per_unit_of_measure = fields.Float(digits=(16, 5), readonly=True)
    inv_sales_amount_expected = fields.Float(digits=(16, 5), readonly=True)

    inv_sales_amount_actual = fields.Float(digits=(16, 5), readonly=True)
    inv_cost_amount_expected = fields.Float(digits=(16, 5), readonly=True)
    inv_cost_amount_non_invtbl = fields.Float(digits=(16, 5), readonly=True)
    inv_item_description = fields.Char(readonly=True)
    inv_cost_amount_expected_acy = fields.Float(digits=(16, 5), readonly=True)
    inv_cost_amount_actual_acy = fields.Float(digits=(16, 5), readonly=True)
    inv_completely_invoiced = fields.Char(readonly=True)
    inv_cost_amount_non_invtbl_acy = fields.Float(digits=(16, 5), readonly=True)
    inv_assemble_to_order = fields.Char(readonly=True)
    inv_drop_shipment = fields.Char(readonly=True)
    inv_open_flag = fields.Char(readonly=True)
    inv_order_type = fields.Char(readonly=True)
    inv_order_no = fields.Char(readonly=True)
    inv_order_line_no = fields.Integer(readonly=True)
    inv_prod_order_comp_line_no = fields.Integer(readonly=True)
    inv_project_no = fields.Char(readonly=True)
    inv_project_task_no = fields.Char(readonly=True)
    inv_source_type = fields.Char(readonly=True)
    inv_source_no = fields.Char(readonly=True)
    inv_source_description = fields.Char(readonly=True)
    inv_source_order_no = fields.Char(readonly=True)
    inv_grade = fields.Char(readonly=True)
    inv_weight = fields.Float(digits=(16, 5), readonly=True)
    inv_posting_date = fields.Date(readonly=True)
    inv_country_of_melt = fields.Char(readonly=True)
    inv_country_of_manufacture = fields.Char(readonly=True)
    inv_source_file = fields.Char(readonly=True)

    mtr_id = fields.Integer(readonly=True)
    mtr_heat_number = fields.Char(readonly=True)
    mtr_batch_number = fields.Char(readonly=True)
    mtr_piece_no = fields.Integer(readonly=True)
    mtr_grade = fields.Char(readonly=True)
    mtr_manufacturer = fields.Char(readonly=True)
    mtr_certificate_number = fields.Char(readonly=True)
    mtr_certificate_date = fields.Date(readonly=True)
    mtr_c = fields.Float(digits=(16, 5), readonly=True)
    mtr_mn = fields.Float(digits=(16, 5), readonly=True)
    mtr_si = fields.Float(digits=(16, 5), readonly=True)
    mtr_p = fields.Float(digits=(16, 5), readonly=True)
    mtr_s = fields.Float(digits=(16, 5), readonly=True)
    mtr_b = fields.Float(digits=(16, 5), readonly=True)
    mtr_v = fields.Float(digits=(16, 5), readonly=True)
    mtr_nb = fields.Float(digits=(16, 5), readonly=True)
    mtr_ti = fields.Float(digits=(16, 5), readonly=True)
    mtr_al = fields.Float(digits=(16, 5), readonly=True)
    mtr_ca = fields.Float(digits=(16, 5), readonly=True)
    mtr_zr = fields.Float(digits=(16, 5), readonly=True)
    mtr_zn = fields.Float(digits=(16, 5), readonly=True)
    mtr_sn = fields.Float(digits=(16, 5), readonly=True)
    mtr_cu = fields.Float(digits=(16, 5), readonly=True)
    mtr_ni = fields.Float(digits=(16, 5), readonly=True)
    mtr_cr = fields.Float(digits=(16, 5), readonly=True)
    mtr_mo = fields.Float(digits=(16, 5), readonly=True)
    mtr_n = fields.Float(digits=(16, 5), readonly=True)
    mtr_ce = fields.Float(digits=(16, 5), readonly=True)
    mtr_yield_strength = fields.Float(digits=(16, 5), readonly=True)
    mtr_tensile_strength = fields.Float(digits=(16, 5), readonly=True)
    mtr_elongation = fields.Float(digits=(16, 5), readonly=True)
    mtr_reduction_area = fields.Float(digits=(16, 5), readonly=True)
    mtr_hardness = fields.Float(digits=(16, 5), readonly=True)
    mtr_thickness = fields.Float(digits=(16, 5), readonly=True)
    mtr_direction = fields.Char(readonly=True)
    mtr_impact_test_temp = fields.Float(digits=(16, 5), readonly=True)
    mtr_impact_charpy = fields.Float(digits=(16, 5), readonly=True)
    mtr_impact_coupon_size = fields.Char(readonly=True)
    mtr_impact_specimen_1 = fields.Float(digits=(16, 5), readonly=True)
    mtr_impact_specimen_2 = fields.Float(digits=(16, 5), readonly=True)
    mtr_impact_specimen_3 = fields.Float(digits=(16, 5), readonly=True)
    mtr_impact_average = fields.Float(digits=(16, 5), readonly=True)
    mtr_country_of_melt = fields.Char(readonly=True)
    mtr_country_of_manufacture = fields.Char(readonly=True)
    mtr_uploaded_at = fields.Datetime(readonly=True)

    def init(self):
        # Drop as regular view OR materialized view depending on what currently exists.
        self.env.cr.execute("""
            DO $$ BEGIN
                IF EXISTS (
                    SELECT 1 FROM pg_class
                    WHERE relname = 'mtr_inventory_join_report' AND relkind = 'v'
                ) THEN
                    DROP VIEW mtr_inventory_join_report CASCADE;
                ELSIF EXISTS (
                    SELECT 1 FROM pg_class
                    WHERE relname = 'mtr_inventory_join_report' AND relkind = 'm'
                ) THEN
                    DROP MATERIALIZED VIEW mtr_inventory_join_report CASCADE;
                END IF;
            END $$
        """)
        self.env.cr.execute(
            """
            CREATE MATERIALIZED VIEW mtr_inventory_join_report AS (
                SELECT
                    i.id AS id,
                    CASE WHEN m.id IS NULL THEN 'Missing MTR' ELSE 'Matched' END AS join_status,
                    CASE WHEN m.id IS NULL THEN 1 ELSE 0 END AS join_priority,
                    i.entry_no AS inv_entry_no,
                    i.date AS inv_date,
                    i.location_code AS inv_location_code,
                    i.item_no AS inv_item_no,
                    i.quantity AS inv_quantity,
                    i.unit_of_measure_code AS inv_unit_of_measure_code,
                    i.document_no AS inv_document_no,
                    i.wsi_variant_code AS inv_wsi_variant_code,
                    i.dimensions AS inv_dimensions,
                    i.lot_number AS inv_lot_number,
                    i.heat_number AS inv_heat_number,
                    i.slab_number AS inv_slab_number,
                    i.internal_bin AS inv_internal_bin,
                    i.additional_notes AS inv_additional_notes,
                    i.cost_amount_actual AS inv_cost_amount_actual,
                    i.description_2 AS inv_description_2,
                    i.origin_code AS inv_origin_code,
                    i.picked AS inv_picked,
                    i.cutting_plan_no AS inv_cutting_plan_no,
                    i.image_path AS inv_image_path,
                    i.entry_type AS inv_entry_type,
                    i.document_type AS inv_document_type,
                    i.drawing AS inv_drawing,
                    i.yield_value AS inv_yield_value,
                    i.document_line_no AS inv_document_line_no,
                    i.revision AS inv_revision,
                    i.laser_quality AS inv_laser_quality,
                    i.unitcost_cwt AS inv_unitcost_cwt,
                    i.piece_no AS inv_piece_no,
                    i.variant_code AS inv_variant_code,
                    i.description AS inv_description,
                    i.return_reason_code AS inv_return_reason_code,
                    i.serial_no AS inv_serial_no,
                    i.package_no AS inv_package_no,
                    i.invoiced_quantity AS inv_invoiced_quantity,
                    i.inventory_by_location AS inv_inventory_by_location,
                    i.inventory AS inv_inventory,
                    i.expiration_date AS inv_expiration_date,
                    i.remaining_quantity AS inv_remaining_quantity,
                    i.shipped_qty_not_returned AS inv_shipped_qty_not_returned,
                    i.reserved_quantity AS inv_reserved_quantity,
                    i.qty_per_unit_of_measure AS inv_qty_per_unit_of_measure,
                    i.sales_amount_expected AS inv_sales_amount_expected,
                    i.sales_amount_actual AS inv_sales_amount_actual,
                    i.cost_amount_expected AS inv_cost_amount_expected,
                    i.cost_amount_non_invtbl AS inv_cost_amount_non_invtbl,
                    i.item_description AS inv_item_description,
                    i.cost_amount_expected_acy AS inv_cost_amount_expected_acy,
                    i.cost_amount_actual_acy AS inv_cost_amount_actual_acy,
                    i.completely_invoiced AS inv_completely_invoiced,
                    i.cost_amount_non_invtbl_acy AS inv_cost_amount_non_invtbl_acy,
                    i.assemble_to_order AS inv_assemble_to_order,
                    i.drop_shipment AS inv_drop_shipment,
                    i.open_flag AS inv_open_flag,
                    i.order_type AS inv_order_type,
                    i.order_no AS inv_order_no,
                    i.order_line_no AS inv_order_line_no,
                    i.prod_order_comp_line_no AS inv_prod_order_comp_line_no,
                    i.project_no AS inv_project_no,
                    i.project_task_no AS inv_project_task_no,
                    i.source_type AS inv_source_type,
                    i.source_no AS inv_source_no,
                    i.source_description AS inv_source_description,
                    i.source_order_no AS inv_source_order_no,
                    i.grade AS inv_grade,
                    i.weight AS inv_weight,
                    i.posting_date AS inv_posting_date,
                    i.country_of_melt AS inv_country_of_melt,
                    i.country_of_manufacture AS inv_country_of_manufacture,
                    i.source_file AS inv_source_file,
                    m.id AS mtr_id,
                    m.heat_number AS mtr_heat_number,
                    m.batch_number AS mtr_batch_number,
                    m.piece_no AS mtr_piece_no,
                    m.grade AS mtr_grade,
                    m.manufacturer AS mtr_manufacturer,
                    m.certificate_number AS mtr_certificate_number,
                    m.certificate_date AS mtr_certificate_date,
                    m.c_element AS mtr_c,
                    m.mn_element AS mtr_mn,
                    m.si_element AS mtr_si,
                    m.p_element AS mtr_p,
                    m.s_element AS mtr_s,
                    m.b_element AS mtr_b,
                    m.v_element AS mtr_v,
                    m.nb_element AS mtr_nb,
                    m.ti_element AS mtr_ti,
                    m.al_element AS mtr_al,
                    m.ca_element AS mtr_ca,
                    m.zr_element AS mtr_zr,
                    m.zn_element AS mtr_zn,
                    m.sn_element AS mtr_sn,
                    m.cu_element AS mtr_cu,
                    m.ni_element AS mtr_ni,
                    m.cr_element AS mtr_cr,
                    m.mo_element AS mtr_mo,
                    m.n_element AS mtr_n,
                    m.ce AS mtr_ce,
                    m.yield_strength AS mtr_yield_strength,
                    m.tensile_strength AS mtr_tensile_strength,
                    m.elongation AS mtr_elongation,
                    m.reduction_area AS mtr_reduction_area,
                    m.hardness AS mtr_hardness,
                    m.thickness AS mtr_thickness,
                    m.direction AS mtr_direction,
                    m.impact_test_temp AS mtr_impact_test_temp,
                    m.impact_charpy AS mtr_impact_charpy,
                    m.impact_coupon_size AS mtr_impact_coupon_size,
                    m.impact_specimen_1 AS mtr_impact_specimen_1,
                    m.impact_specimen_2 AS mtr_impact_specimen_2,
                    m.impact_specimen_3 AS mtr_impact_specimen_3,
                    m.impact_average AS mtr_impact_average,
                    m.country_of_melt AS mtr_country_of_melt,
                    m.country_of_manufacture AS mtr_country_of_manufacture,
                    m.uploaded_at AS mtr_uploaded_at
                FROM inventory i
                LEFT JOIN LATERAL (
                    SELECT m.*
                    FROM mtr_data m
                    WHERE COALESCE(NULLIF(m.certificate_number, ''), '') = COALESCE(NULLIF(i.image_path, ''), '')
                      AND COALESCE(NULLIF(m.batch_number, ''), '') = COALESCE(NULLIF(i.slab_number, ''), '')
                      AND COALESCE(NULLIF(m.heat_number, ''), '') = COALESCE(NULLIF(i.heat_number, ''), '')
                    ORDER BY m.uploaded_at DESC NULLS LAST, m.id DESC
                    LIMIT 1
                ) m ON TRUE
            )
            """
        )
        self.env.cr.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS mtr_inventory_join_report_id_idx "
            "ON mtr_inventory_join_report (id)"
        )
        # Functional indexes on mtr_data so the LATERAL JOIN uses index scans
        # instead of full-table scans, making refresh go from minutes to seconds.
        for idx, expr in [
            ("mtr_data_cert_join_idx",  "COALESCE(NULLIF(certificate_number, ''), '')"),
            ("mtr_data_batch_join_idx", "COALESCE(NULLIF(batch_number, ''), '')"),
            ("mtr_data_heat_join_idx",  "COALESCE(NULLIF(heat_number, ''), '')"),
        ]:
            self.env.cr.execute(
                "CREATE INDEX IF NOT EXISTS %s ON mtr_data (%s)" % (idx, expr)
            )

    @api.model
    def refresh_view(self):
        self.env.cr.execute(
            "REFRESH MATERIALIZED VIEW CONCURRENTLY mtr_inventory_join_report"
        )


class IrModuleModule(models.Model):
    _inherit = "ir.module.module"

    # Compatibility: some website views expect image_ids on ir.module.module.
    image_ids = fields.Many2many(
        "ir.attachment",
        compute="_compute_image_ids",
        string="Images",
        store=False,
    )

    def _compute_image_ids(self):
        empty = self.env["ir.attachment"].browse([])
        for rec in self:
            rec.image_ids = empty
